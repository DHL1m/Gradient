import params
import importlib
import numpy as np
from tensorflow.python.platform import gfile
from funcData import *
from evaluate import evaluate
from tqdm import tqdm
import time
import funcCustom
import funcQuantize
import funcSummary

FLAGS = tf.app.flags.FLAGS

# tf.logging.set_verbosity(FLAGS.log)
if FLAGS.dataset=='cifar10':
    steps=round(50000/FLAGS.batch_size)*20
else:
    steps=round(55000/FLAGS.batch_size)*20
##LR을 decay시켜주는 함수
def learning_rate_decay_fn(learning_rate, global_step,decay_steps=steps):
    print("learning_rate_decay_fn is executed!")
    return tf.train.exponential_decay(
      learning_rate,
      global_step,
      decay_steps=decay_steps,
      decay_rate=0.8,
      staircase=True)
def quantizeGrads(Grad_and_vars,target_level=FLAGS.W_target_level):
    if target_level <= 256*256:
        grads = []
        for grad_and_vars in Grad_and_vars:
            grads.append([funcQuantize.quantize_G(grad_and_vars[0], target_level), grad_and_vars[1]])
        return grads
    return Grad_and_vars

## model을 data로 training 시켜주는 함수
def train(model, data,
          batch_size=128,
          learning_rate=FLAGS.learning_rate,
          Weight_decay=FLAGS.Weight_decay,
          log_dir='./log',
          checkpoint_dir='./checkpoint',
          num_epochs=-1):
    file = open(FLAGS.checkpoint_dir + "/save_model.py", "a")
    with tf.name_scope('data'):
        x, yt = data.next_batch(batch_size)
    global_step =  tf.get_variable('global_step', shape=[],dtype=tf.int64,initializer=tf.constant_initializer(0),
                         trainable=False)
    tf.add_to_collection("Step",global_step)  #Evaluate에서 Drift효과 끄기 위해 구분점역할을 한다.
    y = model(x, is_training=True)
    # Define loss and optimizer
    with tf.name_scope('objective'):
        yt_one=tf.one_hot(yt,10)
        loss = tf.reduce_mean(tf.nn.softmax_cross_entropy_with_logits_v2(labels=yt_one, logits=y), name="loss")
        if Weight_decay:
            loss=loss+tf.reduce_sum(params.WEIGHT_DECAY_FACTOR * tf.stack([tf.nn.l2_loss(i) for i in tf.get_collection('Original_Weight', scope='L')]))
        accuracy=tf.reduce_mean(tf.cast(tf.equal(tf.argmax(yt_one,1), tf.argmax(y, axis=1)),dtype=tf.float32),name="accuracy")
    vars_train=tf.get_collection(tf.GraphKeys.TRAINABLE_VARIABLES,scope='L22')\
               +tf.get_collection(tf.GraphKeys.TRAINABLE_VARIABLES,scope='L23')\
               +tf.get_collection(tf.GraphKeys.TRAINABLE_VARIABLES,scope='L24')\
               +tf.get_collection(tf.GraphKeys.TRAINABLE_VARIABLES,scope='L25') if FLAGS.Fine_tuning else None
    # # * 이런 식으로 gradient  뽑아서  수정가능
    if FLAGS.W_real_target_level<2**10:
        optimizer=tf.train.GradientDescentOptimizer(1)
        grads = optimizer.compute_gradients(loss)
        gradTrainBatch_quantize = quantizeGrads(grads,FLAGS.W_real_target_level)
        opt = optimizer.apply_gradients(gradTrainBatch_quantize, global_step=global_step)
        funcCustom.magic_print("Learning rate =",params.LR_schedule,file=file)
    else:
        opt = tf.contrib.layers.optimize_loss(loss, global_step, learning_rate, optimizer='Adam',
                                              gradient_noise_scale=None, gradient_multipliers=None,
                                              clip_gradients=None, # moving_average_decay=0.9,
                                               update_ops=None, variables=vars_train, name=None)
        funcCustom.magic_print("We use conventional optimization scheme this time, LR =%f"%learning_rate,file=file)


    print("Definite Moving Average...")
    ema = tf.train.ExponentialMovingAverage(params.MOVING_AVERAGE_DECAY, global_step, name='average')
    ema_op = ema.apply([loss, accuracy]+tf.trainable_variables())
    tf.add_to_collection(tf.GraphKeys.UPDATE_OPS, ema_op)
    loss_avg = ema.average(loss)
    tf.summary.scalar('loss/training', loss_avg)
    accuracy_avg = ema.average(accuracy)
    tf.summary.scalar('accuracy/training', accuracy_avg)
    check_loss = tf.check_numerics(loss, 'model diverged: loss->nan')
    tf.add_to_collection(tf.GraphKeys.UPDATE_OPS, check_loss)
    list_W = tf.get_collection('Original_Weight', scope='L')
    list_Wbin = tf.get_collection('Binarized_Weight', scope='L')
    list_Wfluc = tf.get_collection('Fluctuated_Weight', scope='L')
    list_Wread = tf.get_collection('Read_Weight',scope='L')
    list_Wprop = tf.get_collection('Propagated_Weight',scope='L')
    list_Drift_step = tf.get_collection('Drift_step', scope='L')
    list_Drift_value = tf.get_collection('Drift_value', scope='L')
    list_pre_Wbin = tf.get_collection('pre_Wbin', scope='L')
    list_pre_Wfluc = tf.get_collection('pre_Wfluc', scope='L')
    list_pre_Wbin_op = tf.get_collection('pre_Wbin_update_op', scope='L')
    list_pre_Wfluc_op = tf.get_collection('pre_Wfluc_update_op', scope='L')
    list_Lread = tf.get_collection('Level_read',scope='L')
    list_pre_Lread = tf.get_collection('pre_Lread',scope='L')
    list_pre_Lread_op = tf.get_collection('pre_Lread_update_op',scope='L')

    clip_op_list=[]
    with tf.control_dependencies([opt]):
        for ww in list_W:
            clip_op=tf.assign(ww, funcQuantize.clip(ww, FLAGS.W_real_target_level))
            clip_op_list+=[clip_op]
    updates_collection = tf.get_collection(tf.GraphKeys.UPDATE_OPS)
    # updates_collection은 여러 operation으로 이루어진 리스트고, train_op는 그 operation들을 실행하는 operation이다.
    with tf.control_dependencies(clip_op_list):
        train_op = tf.group(*updates_collection)
    print("Make summary for writer...")
    # list_level1 = tf.get_collection("level1")
    # list_level2 = tf.get_collection("level2")
    # list_level3 = tf.get_collection("level3")
    list_prove=tf.get_collection("prove_test")
    if FLAGS.summary:
        funcSummary.add_summaries_scalar([accuracy,loss])
        funcSummary.add_summaries_weight(list_prove)
        # funcSummary.add_summaries_weight(list_W)
        # funcSummary.add_summaries_weight(list_Wbin)

        # funcSummary.add_summaries_weight(list_Wfluc)
        # funcSummary.add_summaries_weight(list_Wread)
        funcSummary.add_summaries_weight(list_Wprop)
        name_list=[]
        for del_W in list_W:  #Use list_W to achieve name of layers
            name_layer=del_W.op.name.split('/')[0]
            name_list.append(name_layer)
            del_collection=tf.get_collection_ref(key='index_history/'+name_layer)
            print("Delete collection:index_history"+name_layer+"=",del_collection)
            del del_collection[:]   #delete collection to avoid error message
    summary_op = tf.summary.merge_all()
    print("Open Session...")
    # Configure options for session
    gpu_options = tf.GPUOptions(allow_growth=True,per_process_gpu_memory_fraction=0.9)
    sess = tf.InteractiveSession(
        config=tf.ConfigProto(
            log_device_placement=False,
            allow_soft_placement=True, gpu_options=gpu_options))
    sess.run(tf.global_variables_initializer())
    for W_ori in list_W:
        W_ori_new=sess.run(tf.assign(W_ori, funcQuantize.quantize_W(W_ori, FLAGS.W_target_level)))
    if FLAGS.Load_checkpoint:
        call_list=tf.get_collection(tf.GraphKeys.VARIABLES)
        print("******We will restore:",call_list)
        saver = tf.train.Saver(max_to_keep=1,var_list=call_list)
        ckpt = tf.train.get_checkpoint_state(params.load_checkpoint_path)
        if ckpt and ckpt.model_checkpoint_path:
            # Restores from checkpoint
            saver.restore(sess, ckpt.model_checkpoint_path)
        else:
            print('No checkpoint file found')
    saver = tf.train.Saver(max_to_keep=1)
    # saver_best= tf.train.Saver(max_to_keep=1)
    summary_writer = tf.summary.FileWriter(log_dir)
    # summary_writer = tf.summary.FileWriter(log_dir, graph=sess.graph)
    coord = tf.train.Coordinator()
    threads = tf.train.start_queue_runners(sess=sess, coord=coord)
    best_acc = 0
    best_loss = 0
    patience = 0
    num_batches = int(data.size[0] / batch_size)
    print("Check the collections...")
    print("list_W:\n",list_W,'\nNum:',len(list_W))
    print("Activations:\n", tf.get_collection(tf.GraphKeys.ACTIVATIONS), '\nNum:', len(tf.get_collection(tf.GraphKeys.ACTIVATIONS)))

    funcCustom.magic_print(
        'We start training..num of trainable param: %d' % funcCustom.count_params(tf.trainable_variables()),file=file)

    for i in range(num_epochs):
        if len(params.LR_schedule)/2:
            if i == params.LR_schedule[0]:
                params.LR_schedule.pop(0)
                LR_new = params.LR_schedule.pop(0)
                if LR_new == 0:
                    print('Optimization Ended!')
                    exit(0)
                LR_old = sess.run(params.LR)
                sess.run(params.LR.assign(LR_new))
                print('lr: %f -> %f' % (LR_old, LR_new))
        if i%10==0 and FLAGS.W_real_target_level<2**12:
            funcCustom.print_statistical_table(sess.run(list_W),name_list=name_list,sheet_name='epochs'+str(i))
        start_time=time.time()
        funcCustom.magic_print('Started epoch %d' % (i + 1), file=file)
        count_num=np.array([0,0,0,0,0,0,0,0,0,0])
        for j in tqdm(range(num_batches)):
            pre_iter_info=sess.run(list_Wbin+list_Wfluc+list_Lread)
            #funcCustom.magic_print(sess.run([list_W]),file=file)
            list_run = sess.run([train_op, loss]+[x,y,yt])
            # if j<2:
            #     funcCustom.plot_images(5,5,list_run[-3],list_run[-1])
            # unique_elements,elements_counts=np.unique(list_run[-1],return_counts=True)
            # num_set=dict(zip(unique_elements,elements_counts))
            # #ii라는 숫자가 dictionary에 들어있다면 카운트에 더해준다.
            # for ii in range(10):
            #     if num_set.__contains__(ii):
            #         count_num[ii]=count_num[ii]+num_set[ii]
            if FLAGS.Inter_variation_options:
                for index, value in enumerate(pre_iter_info[0:len(list_Wbin)]):
                    sess.run(list_pre_Wbin_op[index],{list_pre_Wbin[index]:value})
                for index, value in enumerate(pre_iter_info[len(list_Wbin):len(list_Wbin + list_Wfluc)]):
                    sess.run(list_pre_Wfluc_op[index],{list_pre_Wfluc[index]:value})
                for index, value in enumerate(pre_iter_info[len(list_Wbin + list_Wfluc):len(list_Wbin + list_Wfluc + list_Lread)]):
                    sess.run(list_pre_Lread_op[index],{list_pre_Lread[index]:value})
            if j%100==0:
                summary_writer.add_summary(sess.run(summary_op), global_step=sess.run(global_step))
        step, acc_value, loss_value, summary = sess.run([global_step, accuracy_avg, loss_avg, summary_op])
        # step, acc_value, loss_value = sess.run([global_step, accuracy_avg, loss_avg])
        funcCustom.magic_print(
            ["%d : " % i + str(count_num[i]) for i in range(10)], " Totral num: ", count_num.sum(),file=file)
        funcCustom.magic_print('Training - Accuracy: %.3f' % acc_value, '  Loss:%.3f' % loss_value, file=file)
        saver.save(sess, save_path=checkpoint_dir + '/model.ckpt', global_step=global_step)
        test_acc, test_loss = evaluate(model, FLAGS.dataset,checkpoint_dir=checkpoint_dir)# log_dir=log_dir)
        funcCustom.magic_print('Test     - Accuracy: %.3f' % test_acc, '  Loss:%.3f' % test_loss, file=file)
        print("Elapsed time =",time.time()-start_time)
        if best_acc<test_acc and i!=params.LR_schedule[-2]-1 and i!=num_epochs-1:
            best_acc=test_acc
            best_loss=test_loss
            patience=0
            # saver_best.save(sess, save_path=checkpoint_dir + '/best_model.ckpt', global_step=global_step,latest_filename="best_checkpoint")
        elif best_acc >= test_acc or i==params.LR_schedule[-2]-1 or i==num_epochs-1:
            patience += 1
            if patience > FLAGS.patience or i==params.LR_schedule[-2]-1 or i==num_epochs-1:
                funcCustom.magic_print("Stop this training at epoch" + str(i + 1) + ", because accuracy may be saturated", file=file)
                from openpyxl import Workbook
                from openpyxl import load_workbook
                file_name = 'Accuracy_log'+FLAGS.daystr
                if gfile.Exists(file_name+'.xlsx'):
                    wb=load_workbook(filename=file_name+'.xlsx')
                    ws=wb[file_name]
                else:
                    wb=Workbook()
                    ws=wb.active
                    ws.title=file_name
                row=13 if FLAGS.dataset=='cifar10' else 2  #cifar10:2-10   MNIST:11-19
                row=row+13*(FLAGS.row_choice+FLAGS.choice)
                ws.cell(column=FLAGS.col_choice + 1, row=row, value=FLAGS.memo)
                ws.cell(column=FLAGS.col_choice + 1, row=row + 1, value=FLAGS.W_real_target_level)
                ws.cell(column=FLAGS.col_choice + 1, row=row + 2 , value=FLAGS.W_target_level)
                ws.cell(column=FLAGS.col_choice + 1, row=row + 3, value=FLAGS.Wq_target_level)
                ws.cell(column=FLAGS.col_choice + 1, row=row + 4, value=best_acc)
                ws.cell(column=FLAGS.col_choice + 2, row=row + 4, value=best_loss)
                ws.cell(column=FLAGS.col_choice + 1, row=row + 5, value=test_acc)
                ws.cell(column=FLAGS.col_choice + 2, row=row + 5, value=test_loss)
                ws.cell(column=FLAGS.col_choice + 1, row=row + 6, value=acc_value)
                ws.cell(column=FLAGS.col_choice + 2, row=row + 6, value=loss_value)
                ws.cell(column=FLAGS.col_choice + 1, row=row + 7, value=i)
                ws.cell(column=FLAGS.col_choice + 1, row=row + 8, value="Inter:"+str(params.Inter[0]))
                ws.cell(column=FLAGS.col_choice + 1, row=row + 9, value=str(params.Inter[1:3]))
                ws.cell(column=FLAGS.col_choice + 1, row=row + 10, value=str(params.Inter[3:5]))
                ws.cell(column=FLAGS.col_choice + 2, row=row + 8, value="Intra:"+str(params.Intra[0]))
                ws.cell(column=FLAGS.col_choice + 2, row=row + 9, value=str(params.Intra[1:3]))
                ws.cell(column=FLAGS.col_choice + 2, row=row + 10, value=str(params.Intra[3:5]))
                ws.cell(column=FLAGS.col_choice + 1, row=row + 11, value="Set"+str(FLAGS.choice))
                ws.cell(column=FLAGS.col_choice + 2, row=row + 11, value=FLAGS.dataset+" on "+FLAGS.model)
                wb.save(filename=file_name+'.xlsx')
                break
        funcCustom.magic_print('Best     - Accuracy: %.3f(patience=%d)' % (best_acc, patience), file=file)
        summary_out = tf.Summary()
        summary_out.ParseFromString(summary)
        summary_out.value.add(tag='accuracy/test', simple_value=test_acc)
        summary_out.value.add(tag='loss/test', simple_value=test_loss)
        summary_writer.add_summary(summary_out, step)
        summary_writer.flush()


    # When done, ask the threads to stop.

    file.close()
    coord.request_stop()
    coord.join(threads)
    coord.clear_stop()
    summary_writer.close()
"""
설명2:
1)what is the argv
Argv in Python
The list of command line arguments passed to a Python script. argv[0] is the script name (it is operating system 
dependent whether this is a full pathname or not). If the command was executed using the -c command line option 
to the interpreter, argv[0] is set to the string '-c'.

지금은 FLAGS가 글로벌하게 선언이 되어있어서 argv가 전달된다는게 큰의미는 없다.
(전달안되도 어차피 글로벌이라 그냥 사용가능.tf.app.run()은 그냥 one line fast argument parser로 생각하면 될 듯 하다.)
"""

def main(argv=None):  # pylint: disable=unused-argument
    # 5) 해당 모델을 import한다, importlib.import_module()함수는 코드 과정 중에 패키지를 import 할 때 쓰는 듯 하다.
    if not gfile.Exists(FLAGS.checkpoint_dir):
        # gfile.DeleteRecursively(FLAGS.checkpoint_dir)
        gfile.MakeDirs(FLAGS.checkpoint_dir)
        model_file =( './models/'+FLAGS.model+'.py')
        print(model_file)
        assert gfile.Exists(model_file), 'no model file named: ' + model_file
        gfile.Copy(model_file, FLAGS.checkpoint_dir + '/save_model.py')
        gfile.Copy('./main.py',FLAGS.checkpoint_dir+'/save_main.py')
        gfile.Copy('./evaluate.py',FLAGS.checkpoint_dir+'/save_evaluate.py')
        gfile.Copy('./params.py', FLAGS.checkpoint_dir + '/save_params.py')
        gfile.Copy('./funcData.py', FLAGS.checkpoint_dir + '/save_funcData.py')
        gfile.Copy('./funcCustom.py', FLAGS.checkpoint_dir + '/save_funcCustom.py')
        gfile.Copy('./funcLayer.py', FLAGS.checkpoint_dir + '/save_funcLayer.py')
        gfile.Copy('./funcQuantize.py', FLAGS.checkpoint_dir + '/save_funcQuantize.py')
        gfile.Copy('./funcSummary.py', FLAGS.checkpoint_dir + '/save_funcSummary.py')
    m = importlib.import_module('models.' + FLAGS.model)
    data = get_data_provider(FLAGS.dataset, training=True)
    # data = data_iterator(["./Datasets/MNIST_ori/MNIST_ori_train.tfrecord"],"MNIST")
    # data = data_iterator(["./Datasets/MNIST_backrand/MNIST_backrand_train.tfrecord"], "MNIST")
    # data = data_iterator(["./Datasets/MNIST_back/MNIST_back_train.tfrecord"], "MNIST")
    # data = data_iterator_from_numpy((params.images_train,params.labels_train))
    train(m.model, data,
          batch_size=FLAGS.batch_size,
          checkpoint_dir=FLAGS.checkpoint_dir,
          log_dir=FLAGS.log_dir,
          num_epochs=FLAGS.num_epochs)

"""
#Runs the program with an optional 'main' function and 'argv' list.
->그렇다면 질문은 tf.app.run()가 왜 이렇게 쓰이냐!로 되는데 그건 나중에 생각해보자.
추측: 만약 이 함수를 안쓰면 우리는 string인 cmd 명령어를 쪼개서, 그 중에 어떤 부분이 파일에 정의되어있는지 체크해서
그걸 argv로 만들어줘야한다. 
  nmt_parser = argparse.ArgumentParser()
  add_arguments(nmt_parser)
  FLAGS, unparsed = nmt_parser.parse_known_args()
  tf.app.run(main=main, argv=[sys.argv[0]] + unparsed)        #https://stackoverflow.com/questions/33703624/how-does-tf-app-run-work
보통 위처럼 세줄정도는 나올텐데, 그것보다는 한줄로 써서 위의 기능들을 취하는 것 아닐까?
그리고 main()으로 구분하면 가독성도 좋아지고, __main__이 아니어도 어떻게 접근이 가능할수도 있고..여러가지 가능성이 나오게되는 장점도 있다
"""
if __name__ == '__main__':
    # print(callable(main))
    # print(locals())
    tf.app.run()
